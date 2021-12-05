""" Utility functions for datadumping."""
import os
from typing import Dict, List, Tuple
import re
import datetime
import unicodedata
import csv
import nltk
import pandas
import os
import re
import dateparser
import time
import json
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula import Tokenizer
import functools
import Levenshtein
from icecream import ic


FORMULA_LIST = ['SUM', 'AVERAGE', 'COUNT', 'MAX']


def find_column(coord):
    """ Parse column letter from 'E3'. """
    return re.findall('[a-zA-Z]+', coord)


def find_row(coord):
    """ Parse row number from 'E3'. """
    return re.findall('[0-9]+', coord)


def cell_compare(cell1, cell2):
    """ Compare cell coord by row, then by column."""
    col1, col2 = find_column(cell1)[0], find_column(cell2)[0]
    row1, row2 = find_row(cell1)[0], find_row(cell2)[0]
    if int(row1) < int(row2):
        return -1
    elif int(row1) > int(row2):
        return 1
    else:
        if column_index_from_string(col1) < column_index_from_string(col2):
            return -1
        else:
            return 1


def linked_cell_compare(linked_cell_a, linked_cell_b):
    """ Compare answer cell coord by row, then by column."""
    if isinstance(linked_cell_a[0], str) and isinstance(linked_cell_b[0], str):
        coord_a, coord_b = eval(linked_cell_a[0]), eval(linked_cell_b[0])
    else:
        coord_a, coord_b = linked_cell_a[0], linked_cell_b[0]
    if coord_a[0] < coord_b[0]:
        return -1
    elif coord_a[0] > coord_b[0]:
        return 1
    else:
        if coord_a[1] < coord_b[1]:
            return -1
        else:
            return 1


def sort_region_by_coord(cells):
    """ Sort cells by coords, according to cell_compare(). """
    cell_list = sorted(cells, key=functools.cmp_to_key(cell_compare))
    cell_matrix = []
    last_row = None
    for cell in cell_list:
        col, row = find_column(cell), find_row(cell)
        if row == last_row:
            cell_matrix[-1].append(cell)
        else:
            last_row = row
            cell_matrix.append([cell])
    return cell_list, cell_matrix


def read_annotated(annotated_file_path: str, max_rows: int):
    """ Read annotated xlsx."""
    from preprocess.crawler import Table
    wb = load_workbook(annotated_file_path)

    sheet_names = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_names[0])

    column_dict = {}
    for col in range(1, 11):
        column_dict[col] = ws.cell(1, col).value

    appeared = {}
    table_dict, table_id, record_dict = {}, 0, {}
    for row in range(2, max_rows + 1):
        row_data = {}
        for col in range(1, 11):
            if col == 1:
                row_data[column_dict[col]] = ws.cell(row, col).hyperlink.tooltip
            else:
                row_data[column_dict[col]] = ws.cell(row, col).value

        if (row_data['name&url'], row_data['table id']) not in appeared:
            table_dict[table_id] = Table(table_id, row_data['name&url'], row_data['table id'])
            appeared[(row_data['name&url'], row_data['table id'])] = table_id
            table_id += 1

        record_dict.setdefault(appeared[(row_data['name&url'], row_data['table id'])], []).append(
            dict(
                sentence_id=row,
                subject=row_data['subject'],
                description=row_data['table descriptive sentence id'],
                aggregation=row_data['if has aggregation'] == 'yes',
                data_support=row_data['if data support'] == 'yes'
            )
        )
    return table_dict, record_dict


def load_jsonl(fn):
    result = []
    with open(fn, 'r') as f:
        for line in f:
            data = json.loads(line)
            result.append(data)
    return result


def load_tables(root_path, html_path, valid_html_files=None):
    from preprocess.crawler import Table
    table_dict = {}
    html_dir = os.path.join(root_path, html_path)
    html_files = os.listdir(html_dir)
    if valid_html_files:
        html_files = list(set(valid_html_files) & set(html_files))
    for file_name in html_files:
        # for file_name in ['478.html']:
        file_path = os.path.join(root_path, html_path, file_name)
        with open(file_path, 'r', encoding='utf-8') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
        # table_id = int(file_name.split('.')[0])
        table_id = file_name.split('.')[0]
        table = Table(table_id, None, None)
        table.set_html(soup)
        table_dict[table_id] = table
    return table_dict


def clear_footer(cell):
    footer_links = cell.find_all('a', attrs={'class': ['fn-lnk', 'footnote-link']})
    for footer_link in footer_links:
        footer_link.clear()
    footer_spans = cell.find_all('span', attrs={'class': ['wb-inv', 'wb-invisible']})
    for footer_span in footer_spans:
        footer_span.clear()
    text = cell.text.strip()
    if len(text) < 2:
        return text
    valid_chars = [str(i) for i in range(10)] + [',', '.', '-']
    text = ' '.join(text.split())
    if cell.name == 'td' and ('class' not in cell.attrs or cell['class'] != 'row-heading') \
            and text[-1] not in valid_chars:
        i = len(text) - 2
        while i >= 0 and text[i] == ' ':
            i -= 1
        if i >= 0 and text[i] in valid_chars:
            text = text[:i + 1]
    return text


def auto_fit_column_width(ws):
    for column in ws.columns:
        max_length = 0
        unmerged_cells = list(filter(lambda x: x.coordinate not in ws.merged_cells, column))
        if not unmerged_cells:
            continue
        column_letter = unmerged_cells[0].column_letter  # Get the column name
        for cell in column:
            try:  # Necessary to avoid error on empty cells
                if cell not in unmerged_cells or cell.value is None:
                    continue
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception as e:
                print("Error in when auto-fitting column width: {}".format(e))
                exit()
        adjusted_width = (max_length + 2) * 1.3
        ws.column_dimensions[column_letter].width = adjusted_width


def get_cell_style(cell):
    font = Font()
    fill = PatternFill()
    border_type = Side(border_style='thin', color='00000000')
    border = Border(
        left=border_type,
        right=border_type,
        top=border_type,
        bottom=border_type
    )
    alignment = Alignment(wrap_text=True)
    number_format = 'General'
    # <td>
    if cell.name == 'td':
        # number_format = 'Number'  # TODO: how to automatically set proper format
        if 'class' in cell.attrs and cell['class'][0] == 'row-heading':
            fill.patternType = 'solid'
            fill.fgColor = 'EEEEEEEE'
        if cell.find('strong'):
            font.bold = True
        return font, fill, border, alignment, number_format
    # <th>
    fill.patternType = 'solid'
    fill.fgColor = 'EEEEEEEE'
    alignment.vertical = 'top'
    font.bold = True
    if 'class' not in cell.attrs:
        return font, fill, border, alignment, number_format
    if cell['class'][0] == 'col-left':
        alignment.horizontal = 'left'
    elif cell['class'][0] == 'col-group':
        pass
        # alignment.horizontal = 'center'
    elif cell['class'][0] == 'row-heading':
        pass
    elif cell['class'][0] == 'row-stub':
        font.bold = False
        fill.patternType = None
    elif re.match('stub-indent\d', cell['class'][0]):
        font.bold = False
        fill.patternType = None
        indent_start_idx = len('stub-indent')
        alignment.indent = int(cell['class'][0][indent_start_idx:])
    elif re.match('heading-indent\d', cell['class'][0]):
        indent_start_idx = len('heading-indent')
        alignment.indent = int(cell['class'][0][indent_start_idx:])
    elif cell['class'][0] == 'uom-center':
        font.bold = False
        alignment.horizontal = 'center'
    elif cell['class'][0] == 'uom-left':
        font.bold = False
        alignment.horizontal = 'left'
    elif cell['class'][0] == 'uom':
        font.bold = False

    if cell.find('strong'):
        font.bold = True
    return font, fill, border, alignment, number_format


def prettify_caption(caption):
    for i in range(len(caption)):
        caption[i] = caption[i].strip()
    try:
        summary_idx = caption.index('Table summary')
        return f"{': '.join(caption[:summary_idx])}\n" \
               f"{': '.join(caption[summary_idx:])}"
    except:
        return ' '.join(caption)


def prettify_title(caption):
    for i in range(len(caption)):
        caption[i] = caption[i].strip()
    try:
        summary_idx = caption.index('Table summary')
        if summary_idx >= 1:
            return caption[summary_idx - 1]
        else:
            return ' '.join(caption[summary_idx + 1:])
    except:
        return caption[-1]


def find_ranges(table):
    max_rows = len(table.html.find_all('tr'))
    footer_rows = len(table.html.find_all('td', attrs={'class': re.compile('table-footer')}))
    max_rows -= footer_rows

    max_cols = 0
    first_row = table.html.find('tr')
    for cell in first_row.find_all(['th', 'td']):
        colspan = get_int(cell, 'colspan')
        max_cols += colspan
    return max_rows, max_cols


def find_position_zeroindex(table_mask, row, col, max_rows, max_cols):
    while table_mask[row, col] == 1:
        col += 1
        if col > max_cols:
            col = 0
            row += 1
            if row >= max_rows:
                raise ValueError("Can not find position for row {}".format(row))
    return row, col


def find_position(table_mask, row, col, max_rows, max_cols):
    while table_mask[row, col] == 1:
        col += 1
        if col > max_cols:
            col = 1
            row += 1
            if row > max_rows:
                raise ValueError("Can not find position for row {}".format(row))
    return row, col


def get_int(cell, key):
    try:
        return int(cell.get(key, 1))
    except ValueError:
        try:
            return int(re.search('[0-9]+', cell[key]).group())
        except:
            return 1


# --------------------------------------------
# Normalize and Inferring Types.
def normalize(x):
    """ Normalize header string. """
    # Copied from WikiTableQuestions dataset official evaluator.
    if x is None:
        return None
    # Remove diacritics
    x = ''.join(c for c in unicodedata.normalize('NFKD', x)
                if unicodedata.category(c) != 'Mn')
    # Normalize quotes and dashes
    x = re.sub("[‘’´`]", "'", x)
    x = re.sub("[“”]", "\"", x)
    x = re.sub("[‐‑‒–—−]", "-", x)
    while True:
        old_x = x
        # Remove citations
        x = re.sub("((?<!^)\[[^\]]*\]|\[\d+\]|[•♦†‡*#+])*$", "", x.strip())
        # Remove details in parenthesis
        x = re.sub("(?<!^)( \([^)]*\))*$", "", x.strip())
        # Remove outermost quotation mark
        x = re.sub('^"([^"]*)"$', r'\1', x.strip())
        if x == old_x:
            break
    # Remove final '.'
    if x and x[-1] == '.':
        x = x[:-1]
    # Collapse whitespaces and convert to lower case
    x = re.sub('\s+', ' ', x, flags=re.U).lower().strip()
    return x


def naive_str_to_float(string):
    """ A naive way to convert str to float, if convertable."""
    sanitized = string
    try:
        if sanitized[0] == '(':
            sanitized = sanitized[1:]
        if (sanitized[-1] == '%') or (sanitized[-1] == ')'):
            sanitized = sanitized[: -1]
        sanitized = sanitized.replace(',', '')
        new = float(sanitized)
        return new
    except:
        return normalize(string)


def infer_type(name: str, client):
    """ Infer real type of header cell by stanza/dateparser."""
    # annotate by corenlp
    annotation = client.annotate(name)
    value, type = name, 'string'

    if len(annotation.sentence) > 1 or len(annotation.sentence[0].token) > 10:
        return value, type

    # get NER str
    ner_str, ner_type = find_ner_str(name, annotation)

    try:
        if ner_type == 'DURATION':  # if 'DURATION', parse by hand.
            date_groups = re.findall('\d+', ner_str)
            start_date, end_date, *others = date_groups
            start_date_str = dateparser.parse(start_date, settings={'RELATIVE_BASE': datetime.datetime(2000, 1, 1)}) \
                .strftime("%Y-%m-%d")
            end_date_str = dateparser.parse(end_date, settings={'RELATIVE_BASE': datetime.datetime(2000, 1, 1)}) \
                .strftime("%Y-%m-%d")
            value = start_date_str + '<TO>' + end_date_str
            type = 'datetime'
            print(f"{ner_type}: {name} => {value}")
        elif ner_type == 'DATE':  # if 'DATE', parse by dateparser.
            if re.search('(to|T[o|O])', ner_str):
                start_date, end_date, *others = re.findall('\d+', ner_str)
                start_date_str = dateparser.parse(start_date, settings={'RELATIVE_BASE': datetime.datetime(2000, 1, 1)}) \
                    .strftime("%Y-%m-%d")
                end_date_str = dateparser.parse(end_date, settings={'RELATIVE_BASE': datetime.datetime(2000, 1, 1)}) \
                    .strftime("%Y-%m-%d")
                value = start_date_str + '<TO>' + end_date_str
            else:
                value = dateparser.parse(ner_str, settings={'RELATIVE_BASE': datetime.datetime(2000, 1, 1)}) \
                    .strftime("%Y-%m-%d")
            type = 'datetime'
            print(f"{ner_type}: {name} => {value}")
        elif ner_type == 'NUMBER':  # if 'NUMBER', check if number range and parse by hand
            if re.fullmatch('^\d+\s*-?\s*(to|T[o|O])?\s*-?\s*\d+$', ner_str):  # exclude 2017/18, 2017/1/12
                num_groups = re.findall('\d+', ner_str)
                if len(num_groups) > 1:  # number range
                    start_num_str, end_num_str, *others = num_groups
                    if float(end_num_str) > float(start_num_str):  # exclude 2017-18
                        value = start_num_str + '<TO>' + end_num_str
                        type = 'number'
                else:  # number
                    value = float(num_groups[0])
                    type = 'number'
                if type == 'number':
                    print(f"{ner_type}: {name} => {value}")
                else:
                    print(f"Failed to infer {name} to {ner_type}")
    except Exception as e:
        print(f"In infer_type(): {e}")

    return value, type


def find_ner_str(name: str, annotation):
    """ Find the NER substring from name."""
    type, curr_type = None, None
    start_idx, end_idx = -1, -1
    for token in annotation.sentence[0].token:
        if token.timexValue.type:
            curr_type = token.timexValue.type
        elif token.ner in ['DURATION', 'DATE', 'NUMBER']:
            curr_type = token.ner

        if type is None and curr_type is not None:  # the first Date/Number token
            type = curr_type
            start_idx = token.beginChar
            end_idx = token.endChar
        elif type is not None and curr_type == type:
            end_idx = token.endChar

    if type is None:
        return None, 'string'
    else:
        return name[start_idx: end_idx], type


def find_aggr_by_formula(sheet, top_root, left_root, aggr_coords: Dict):
    """ Find aggregation row/column by excel formula."""
    top_indices = top_root.take_indices("top")
    left_indices = left_root.take_indices("left")

    for i, iname in left_indices:
        for j, jname in top_indices:
            cell_value = sheet[i][j - 1].value
            if isinstance(cell_value, str) and is_formula(cell_value):
                dir = get_aggr_direction(cell_value)
                if dir == 'top':
                    aggr_coords[dir].add(idx2str(j))  # TODO: j or j-1?
                else:
                    aggr_coords[dir].add(str(i))


def get_aggr_direction(cell_value: str) -> str:
    """ Get the direction of aggregation."""
    start_idx, end_idx = cell_value.find('('), cell_value.find(')')
    cell_value = cell_value[start_idx + 1: end_idx + 2]
    # start_coord, end_coord = cell_value.split(':')
    start_coord, end_coord, *others = re.findall('[a-zA-Z]+[0-9]+', cell_value)
    start_coord_top = re.findall('[a-zA-Z]+', start_coord)[0]
    end_coord_top = re.findall('[a-zA-Z]+', end_coord)[0]

    if start_coord_top == end_coord_top:
        return 'left'
    else:
        return 'top'


def is_formula(cell_value: str):
    """ Check if a cell_value is excel formula."""
    return cell_value.split('(')[0][1:] in FORMULA_LIST


def idx2str(col_index: str):
    """ Convert numeric column index to excel column index."""
    column_str = ""
    column_remainder = col_index
    column_base = ord("A")
    while column_remainder > 0:
        column_remainder -= 1
        new = (column_remainder % 26) + column_base
        column_str = chr(new) + column_str
        column_remainder = column_remainder // 26
    return column_str
